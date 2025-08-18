"""
Adds Teamnames, times and gamenumber to GameReport
"""
import openpyxl, os

from SwissTournament import SwissTournament

class Report:

    def __init__(self, game, number, dates, templatePath):
        self.__date = dates[game.day]
        self.__teamA = game.teamA
        self.__teamB = game.teamB
        self.__referee = game.referee.name
        self.__time = game.time
        self.__templatePath = templatePath
        self.__gameNumber = number
        filename = "report_Game-" + str(number) + ".xlsx"
        pdfFileName = "report_Game-" + str(number) + ".pdf"
        self.__outputPath = os.path.join(os.path.dirname(templatePath), filename)

    def generateReport(self,  teamNames):
        template = openpyxl.load_workbook(self.__templatePath)
        sheet = template.active 
        sheet.cell(row=7, column=3).value = self.__time
        sheet.cell(row=10, column=1).value = self.__date
        sheet.cell(row=7, column=1).value = self.__gameNumber
        if teamNames and "Platz" not in self.__teamA.name and "Platz" not in self.__teamB.name:
            sheet.cell(row=12, column=4).value = self.__referee
            sheet.cell(row=22, column=2).value = self.__teamA.name
            sheet.cell(row=22, column=6).value = self.__teamB.name
            self.__enterPlayers(sheet, self.__teamA.players, [24, [1, 2, 4]])
            self.__enterPlayers(sheet, self.__teamB.players, [24, [5, 6, 7]])
        template.save(filename=self.__outputPath)

    def __enterPlayers(self, sheet, players, startCell):
        currentCell = startCell
        for player in players:
            if player.number > 0:
                sheet.cell(row=currentCell[0], column=currentCell[1][0]).value = player.number
                sheet.cell(row=currentCell[0], column=currentCell[1][1]).value = player.name[0]
                sheet.cell(row=currentCell[0], column=currentCell[1][2]).value = player.name[1]
                currentCell[0] += 1

class GenerateReports:

    templatePath = os.path.join(os.getcwd(), "Reports", "template.xlsx")

    @staticmethod
    def generateGroupReports(tournament, teamNames):
        for index, game in enumerate(tournament.games):
            try:
                report = Report(game, index+1, tournament.dates, GenerateReports.templatePath)
                report.generateReport(teamNames)
            except:
                print("Error occured stopped at Game Nr. " + str(index+1))
                break
            

if __name__ == "__main__":
    tournament = SwissTournament()
    GenerateReports.generateGroupReports(tournament, teamNames=True)
